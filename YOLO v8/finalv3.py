import pandas as pd
import time
from datetime import datetime
import cv2
from ultralytics import YOLO
import time
import pandas as pd
from openpyxl import load_workbook
URL = 0
excel_file_path = "object_detection.xlsx"
model_path = "D:/Vibehai/IPD/YOLO v8/best3.pt"
model = YOLO(model_path)
cap = cv2.VideoCapture(URL)
class_list = ['Instagram', 'None','Book']
current_datetime = datetime.now()

# Format the date as "dd mm yy"
formatted_date = current_datetime.strftime("%d %m %y")

# Format the time as "hh min sec"
formatted_time = current_datetime.strftime("%H %M %S")


def pred():

    start_time = time.time()  # Get the current time in seconds
    end_time = start_time + 3  # Calculate the end time (3 seconds from the start)
    list = []
    list_dict = []
    while time.time() < end_time:
        ret, frame = cap.read()

        if not ret:
            print("Obstruction in Input")

        # Perform object detection using YOLO
        results = model(frame)

        a=results[0].boxes.data
        px=pd.DataFrame(a).astype("float")
    #   
        for _,row in px.iterrows():
            current_datetime = datetime.now()

            # Format the date as "dd mm yy"
            formatted_date = current_datetime.strftime("%d %m %y")

            # Format the time as "hh min sec"
            formatted_time = current_datetime.strftime("%H %M %S")

    #       print(row)
            d=int(row[5])
            c=class_list[d]
            z = {"Task": c, "Time": formatted_time, "Date": formatted_date} #where to add
            if c:
                list_dict.append(z)

    if list_dict:
        return list_dict
    else:
        return "No Detections"
    

data_log = pd.DataFrame(columns=["Time", "Returned String"])

try:
    while True:
        #Every 5 minutes ////////
        time.sleep(3)
        returned_dict = pred()
        if returned_dict == 'No Detections':
            current_datetime = datetime.now()

            formatted_date = current_datetime.strftime("%d %m %y")

            formatted_time = current_datetime.strftime("%H %M %S")

            save_dict = [{"Task": "No Detections","Time": formatted_time, "Date": formatted_date}]
        else:

            save_dict = returned_dict

     # Append data to the existing DataFrame
        # data_log = data_log.append(save_dict, ignore_index=True)
        
        # Save the DataFrame to the Excel file
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', datetime_format='hh:mm:ss') as writer:
            data_log.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Format "Time" and "Date" columns
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active

        # Get the dimensions of the data
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Apply formatting to the "Time" and "Date" columns
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=3):
            for cell in row:
                if cell.column == 2:  # "Time" column
                    cell.number_format = 'hh:mm:ss'
                elif cell.column == 3:  # "Date" column
                    cell.number_format = 'dd mm yy'

        # Save the formatted Excel file
        workbook.save(excel_file_path)
except KeyboardInterrupt:
    print("Data logging stopped.")